"""ESG Excel 导出器模块

提供将ESG指标数据、分析结果导出为Excel文件的功能。
支持多工作表导出、格式化和图表生成。
"""

import logging
import os
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

# 配置日志
logger = logging.getLogger(__name__)

# 尝试导入 openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    Workbook = None

# 尝试导入 pandas
try:
    import pandas as pd

    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    pd = None


@dataclass
class ExportConfig:
    """导出配置类

    Attributes:
        include_scores: 是否包含评分
        include_charts: 是否包含图表
        include_benchmark: 是否包含基准对比
        sheet_names: 工作表名称配置
        output_dir: 输出目录
        filename_prefix: 文件名前缀
    """

    include_scores: bool = True
    include_charts: bool = True
    include_benchmark: bool = True
    sheet_names: Dict[str, str] = field(
        default_factory=lambda: {
            "summary": "ESG概览",
            "environment": "环境指标(E)",
            "social": "社会指标(S)",
            "governance": "治理指标(G)",
            "scores": "评分详情",
            "benchmark": "基准对比",
        }
    )
    output_dir: str = "./output"
    filename_prefix: str = "ESG_Report"


class ExcelExportError(Exception):
    """Excel 导出异常基类"""

    pass


class LibraryNotFoundError(ExcelExportError):
    """依赖库未安装异常"""

    pass


class ExcelExporter:
    """ESG Excel 导出器

    将ESG指标数据和分析结果导出为格式化的Excel文件。

    Example:
        >>> from src.esg.core.models import ESGMetrics
        >>> metrics = ESGMetrics(company_name="示例公司", year="2024")
        >>> exporter = ExcelExporter()
        >>> exporter.export(metrics, "report.xlsx")
    """

    # 样式配置
    HEADER_FILL = None
    HEADER_FONT = None
    BORDER_STYLE = None
    CENTER_ALIGNMENT = None

    def __init__(self, config: Optional[ExportConfig] = None) -> None:
        """初始化 Excel 导出器

        Args:
            config: 导出配置，为 None 时使用默认配置
        """
        self.config = config or ExportConfig()
        self._validate_libraries()
        self._init_styles()

    def _validate_libraries(self) -> None:
        """验证必要的库是否已安装"""
        if not HAS_OPENPYXL:
            raise LibraryNotFoundError(
                "未找到 openpyxl 库，请安装：pip install openpyxl"
            )

    def _init_styles(self) -> None:
        """初始化 Excel 样式"""
        if HAS_OPENPYXL:
            # 表头样式
            self.HEADER_FILL = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            self.HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

            # 边框样式
            self.BORDER_STYLE = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # 居中对齐
            self.CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    def export(
        self,
        metrics: Any,
        output_path: Union[str, Path],
        analysis_result: Optional[Any] = None,
        benchmark_data: Optional[Any] = None,
    ) -> Path:
        """导出 ESG 指标到 Excel 文件

        Args:
            metrics: ESGMetrics 对象
            output_path: 输出文件路径
            analysis_result: 可选的 AnalysisResult 对象
            benchmark_data: 可选的 BenchmarkData 对象

        Returns:
            Path: 实际保存的文件路径

        Raises:
            ExcelExportError: 导出失败时
        """
        output = Path(output_path)

        # 确保输出目录存在
        output.parent.mkdir(parents=True, exist_ok=True)

        try:
            workbook = Workbook()
            # 移除默认工作表
            default_sheet = workbook.active
            if default_sheet:
                workbook.remove(default_sheet)

            # 创建概览工作表
            self._create_summary_sheet(workbook, metrics, analysis_result)

            # 创建环境指标工作表
            self._create_dimension_sheet(workbook, metrics, "E")

            # 创建社会指标工作表
            self._create_dimension_sheet(workbook, metrics, "S")

            # 创建治理指标工作表
            self._create_dimension_sheet(workbook, metrics, "G")

            # 创建评分详情工作表
            if self.config.include_scores:
                self._create_scores_sheet(workbook, metrics, analysis_result)

            # 创建基准对比工作表
            if self.config.include_benchmark and benchmark_data:
                self._create_benchmark_sheet(workbook, metrics, benchmark_data)

            # 保存文件
            workbook.save(output)
            logger.info(f"ESG报告已导出: {output}")

            return output

        except Exception as e:
            raise ExcelExportError(f"Excel导出失败: {e}")

    def _create_summary_sheet(
        self, workbook: Any, metrics: Any, analysis_result: Optional[Any]
    ) -> None:
        """创建概览工作表

        Args:
            workbook: Excel 工作簿对象
            metrics: ESGMetrics 对象
            analysis_result: 可选的 AnalysisResult 对象
        """
        sheet_name = self.config.sheet_names.get("summary", "ESG概览")
        sheet = workbook.create_sheet(title=sheet_name)

        # 公司基本信息
        summary_data = [
            ["ESG 分析报告"],
            [],
            ["公司名称", metrics.company_name],
            ["报告年份", metrics.year],
            ["行业类型", metrics.industry_sector],
            ["数据来源", metrics.source or "PDF提取"],
            ["提取时间", metrics.extracted_at],
            [],
            ["ESG 评分概览"],
            ["环境 (E)", metrics.get_dimension_score("E")],
            ["社会 (S)", metrics.get_dimension_score("S")],
            ["治理 (G)", metrics.get_dimension_score("G")],
            ["综合得分", self._calculate_overall_score(metrics)],
            [],
            ["置信度等级", metrics.calculate_overall_confidence()],
        ]

        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=16)
                elif row_idx in [3, 4, 5, 6, 7, 10, 11, 12, 13, 16]:
                    if col_idx == 1:
                        cell.font = Font(bold=True)

        # 调整列宽
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 30

    def _create_dimension_sheet(
        self, workbook: Any, metrics: Any, dimension: str
    ) -> None:
        """创建维度指标工作表

        Args:
            workbook: Excel 工作簿对象
            metrics: ESGMetrics 对象
            dimension: 维度代码 ('E', 'S', 'G')
        """
        sheet_names = {
            "E": self.config.sheet_names.get("environment", "环境指标(E)"),
            "S": self.config.sheet_names.get("social", "社会指标(S)"),
            "G": self.config.sheet_names.get("governance", "治理指标(G)"),
        }
        sheet = workbook.create_sheet(title=sheet_names.get(dimension, f"维度{dimension}"))

        # 获取维度指标定义
        indicators = self._get_dimension_indicators(metrics, dimension)

        # 表头
        headers = ["指标名称", "指标值", "单位", "数据来源", "置信度"]
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER_STYLE
            cell.alignment = self.CENTER_ALIGNMENT

        # 数据行
        for row_idx, (name, value, unit, source, confidence) in enumerate(
            indicators, 2
        ):
            row_data = [name, value, unit, source, confidence]
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.border = self.BORDER_STYLE

        # 调整列宽
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 10
        sheet.column_dimensions["D"].width = 20
        sheet.column_dimensions["E"].width = 10

        # 维度得分
        score_row = len(indicators) + 3
        sheet.cell(row=score_row, column=1, value="维度得分").font = Font(bold=True)
        sheet.cell(row=score_row, column=2, value=metrics.get_dimension_score(dimension))

    def _get_dimension_indicators(
        self, metrics: Any, dimension: str
    ) -> List[tuple]:
        """获取维度指标列表

        Args:
            metrics: ESGMetrics 对象
            dimension: 维度代码

        Returns:
            List[tuple]: 指标元组列表 (名称, 值, 单位, 来源, 置信度)
        """
        indicators = []

        if dimension == "E":
            # 环境指标
            e_indicators = [
                ("碳排放总量", metrics.carbon_emissions, "吨", metrics.data_sources.get("carbon_emissions", "")),
                ("范围1排放", metrics.scope1_emissions, "吨", metrics.data_sources.get("scope1_emissions", "")),
                ("范围2排放(位置法)", metrics.scope2_emissions_location, "吨", metrics.data_sources.get("scope2_emissions_location", "")),
                ("范围2排放(市场法)", metrics.scope2_emissions_market, "吨", metrics.data_sources.get("scope2_emissions_market", "")),
                ("范围3排放", metrics.scope3_emissions, "吨", metrics.data_sources.get("scope3_emissions", "")),
                ("碳强度", metrics.carbon_intensity, "吨CO2e/百万元", metrics.data_sources.get("carbon_intensity", "")),
                ("可再生能源比例", metrics.renewable_energy_ratio, "%", metrics.data_sources.get("renewable_energy_ratio", "")),
                ("能源效率", metrics.energy_efficiency, "", metrics.data_sources.get("energy_efficiency", "")),
                ("用水量", metrics.water_consumption, "立方米", metrics.data_sources.get("water_consumption", "")),
                ("水资源强度", metrics.water_intensity, "立方米/百万元", metrics.data_sources.get("water_intensity", "")),
                ("废物回收率", metrics.waste_recycling_rate, "%", metrics.data_sources.get("waste_recycling_rate", "")),
                ("生物多样性评分", metrics.biodiversity_impact_score, "分", metrics.data_sources.get("biodiversity_impact_score", "")),
                # 新能源特色指标
                ("风机可利用率", metrics.turbine_availability, "%", metrics.data_sources.get("turbine_availability", "")),
                ("弃风弃光率", metrics.curtailment_rate, "%", metrics.data_sources.get("curtailment_rate", "")),
                ("电池循环寿命", metrics.battery_cycle_life, "次", metrics.data_sources.get("battery_cycle_life", "")),
                ("电池回收率", metrics.battery_recycling_rate, "%", metrics.data_sources.get("battery_recycling_rate", "")),
                ("电解效率", metrics.electrolysis_efficiency, "%", metrics.data_sources.get("electrolysis_efficiency", "")),
                ("储能安全评分", metrics.energy_storage_safety_score, "分", metrics.data_sources.get("energy_storage_safety_score", "")),
            ]
            indicators = [(n, v, u, s, metrics.confidence.get(n, 0)) for n, v, u, s in e_indicators if v is not None]

        elif dimension == "S":
            # 社会指标
            s_indicators = [
                ("员工数量", metrics.employee_count, "人", metrics.data_sources.get("employee_count", "")),
                ("女性员工比例", metrics.female_ratio, "%", metrics.data_sources.get("female_ratio", "")),
                ("高管女性比例", metrics.female_executive_ratio, "%", metrics.data_sources.get("female_executive_ratio", "")),
                ("人均培训时长", metrics.training_hours, "小时", metrics.data_sources.get("training_hours", "")),
                ("人均培训投入", metrics.training_investment_per_employee, "美元", metrics.data_sources.get("training_investment_per_employee", "")),
                ("安全事故数量", metrics.safety_incidents, "起", metrics.data_sources.get("safety_incidents", "")),
                ("TRIR", metrics.trir, "", metrics.data_sources.get("trir", "")),
                ("LTIFR", metrics.ltifr, "", metrics.data_sources.get("ltifr", "")),
                ("安全投入占比", metrics.safety_investment_ratio, "%", metrics.data_sources.get("safety_investment_ratio", "")),
                ("本地雇佣比例", metrics.local_employment_ratio, "%", metrics.data_sources.get("local_employment_ratio", "")),
                ("社区投资金额", metrics.community_investment, "元", metrics.data_sources.get("community_investment", "")),
                ("社区投资占营收比例", metrics.community_investment_per_revenue, "%", metrics.data_sources.get("community_investment_per_revenue", "")),
            ]
            indicators = [(n, v, u, s, metrics.confidence.get(n, 0)) for n, v, u, s in s_indicators if v is not None]

        elif dimension == "G":
            # 治理指标
            g_indicators = [
                ("董事会独立比例", metrics.board_independence_ratio, "%", metrics.data_sources.get("board_independence_ratio", "")),
                ("道德培训覆盖率", metrics.ethics_training_coverage, "%", metrics.data_sources.get("ethics_training_coverage", "")),
                ("ESG报告质量", metrics.esg_report_quality, "分", metrics.data_sources.get("esg_report_quality", "")),
                ("ESG委员会独立性", metrics.esg_committee_independence, "", metrics.data_sources.get("esg_committee_independence", "")),
                ("反腐败培训覆盖率", metrics.anti_corruption_training_coverage, "%", metrics.data_sources.get("anti_corruption_training_coverage", "")),
                ("举报人保护机制", "是" if metrics.whistleblower_protection else "否" if metrics.whistleblower_protection is not None else "", "", metrics.data_sources.get("whistleblower_protection", "")),
            ]
            indicators = [(n, v, u, s, metrics.confidence.get(n, 0)) for n, v, u, s in g_indicators if v is not None and v != ""]

        return indicators

    def _create_scores_sheet(
        self, workbook: Any, metrics: Any, analysis_result: Optional[Any]
    ) -> None:
        """创建评分详情工作表

        Args:
            workbook: Excel 工作簿对象
            metrics: ESGMetrics 对象
            analysis_result: 可选的 AnalysisResult 对象
        """
        sheet_name = self.config.sheet_names.get("scores", "评分详情")
        sheet = workbook.create_sheet(title=sheet_name)

        # 评分数据
        scores = metrics.get_all_dimension_scores()

        # 表头
        headers = ["维度", "得分", "评级", "说明"]
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER_STYLE

        # 维度评分数据
        dimension_info = {
            "E": ("环境", "环境维度评估企业的碳排放、能源效率、资源利用等表现"),
            "S": ("社会", "社会维度评估企业的员工权益、安全健康、社区责任等表现"),
            "G": ("治理", "治理维度评估企业的治理结构、商业道德、信息披露等表现"),
        }

        row_idx = 2
        for dim, score in scores.items():
            name, desc = dimension_info.get(dim, (dim, ""))
            rating = self._get_score_rating(score)
            row_data = [name, score, rating, desc]
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER_STYLE

                # 根据评分设置颜色
                if col_idx == 2:
                    cell.fill = self._get_score_fill(score)

            row_idx += 1

        # 综合得分
        overall_score = self._calculate_overall_score(metrics)
        overall_row = row_idx + 1
        sheet.cell(row=overall_row, column=1, value="综合得分").font = Font(bold=True)
        sheet.cell(row=overall_row, column=2, value=overall_score)
        sheet.cell(row=overall_row, column=3, value=self._get_score_rating(overall_score))

        # 调整列宽
        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 10
        sheet.column_dimensions["C"].width = 10
        sheet.column_dimensions["D"].width = 50

        # 添加图表
        if self.config.include_charts:
            self._add_score_chart(sheet, scores, row_idx + 3)

    def _create_benchmark_sheet(
        self, workbook: Any, metrics: Any, benchmark_data: Any
    ) -> None:
        """创建基准对比工作表

        Args:
            workbook: Excel 工作簿对象
            metrics: ESGMetrics 对象
            benchmark_data: BenchmarkData 对象
        """
        sheet_name = self.config.sheet_names.get("benchmark", "基准对比")
        sheet = workbook.create_sheet(title=sheet_name)

        # 表头
        headers = ["指标", "企业值", "行业基准", "差距", "差距率(%)"]
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER_STYLE

        # 对比数据
        comparisons = [
            ("可再生能源比例(%)", metrics.renewable_energy_ratio, benchmark_data.avg_renewable_energy_ratio),
            ("能源效率", metrics.energy_efficiency, benchmark_data.avg_energy_efficiency),
            ("女性员工比例(%)", metrics.female_ratio, benchmark_data.avg_female_ratio),
            ("培训时长(小时)", metrics.training_hours, benchmark_data.avg_training_hours),
            ("董事会独立比例(%)", metrics.board_independence_ratio, benchmark_data.avg_board_independence_ratio),
            ("ESG报告质量(分)", metrics.esg_report_quality, benchmark_data.avg_esg_report_quality),
            ("碳强度", metrics.carbon_intensity, benchmark_data.avg_carbon_intensity),
            ("水资源强度", metrics.water_intensity, benchmark_data.avg_water_intensity),
        ]

        row_idx = 2
        for name, company_val, benchmark_val in comparisons:
            if company_val is not None and benchmark_val is not None:
                gap = company_val - benchmark_val
                gap_pct = (gap / benchmark_val * 100) if benchmark_val != 0 else 0

                row_data = [name, company_val, benchmark_val, gap, round(gap_pct, 2)]
                for col_idx, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.BORDER_STYLE

                    # 差距列设置条件颜色
                    if col_idx == 4:
                        if gap > 0:
                            cell.fill = PatternFill(
                                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                            )
                        elif gap < 0:
                            cell.fill = PatternFill(
                                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                            )

                row_idx += 1

        # 基准信息
        info_row = row_idx + 2
        sheet.cell(row=info_row, column=1, value=f"基准行业: {benchmark_data.industry}")
        sheet.cell(row=info_row + 1, column=1, value=f"基准年份: {benchmark_data.year}")
        sheet.cell(row=info_row + 2, column=1, value=f"样本数量: {benchmark_data.sample_size}")
        sheet.cell(row=info_row + 3, column=1, value=f"数据来源: {benchmark_data.source}")

        # 调整列宽
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 12
        sheet.column_dimensions["D"].width = 10
        sheet.column_dimensions["E"].width = 12

    def _add_score_chart(
        self, sheet: Any, scores: Dict[str, float], start_row: int
    ) -> None:
        """添加评分柱状图

        Args:
            sheet: Excel 工作表对象
            scores: 各维度得分
            start_row: 图表起始行
        """
        # 在工作表中添加图表数据
        sheet.cell(row=start_row, column=1, value="维度")
        sheet.cell(row=start_row, column=2, value="得分")

        dimension_names = {"E": "环境", "S": "社会", "G": "治理"}
        for idx, (dim, score) in enumerate(scores.items(), start_row + 1):
            sheet.cell(row=idx, column=1, value=dimension_names.get(dim, dim))
            sheet.cell(row=idx, column=2, value=score)

        # 创建柱状图
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "ESG 各维度得分"
        chart.y_axis.title = "得分"
        chart.x_axis.title = "维度"

        data = Reference(sheet, min_col=2, min_row=start_row, max_row=start_row + 3)
        categories = Reference(sheet, min_col=1, min_row=start_row + 1, max_row=start_row + 3)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.shape = 4
        chart.width = 12
        chart.height = 8

        sheet.add_chart(chart, "F2")

    def _calculate_overall_score(self, metrics: Any) -> float:
        """计算综合得分

        Args:
            metrics: ESGMetrics 对象

        Returns:
            float: 综合得分
        """
        scores = metrics.get_all_dimension_scores()
        weights = {"E": 0.4, "S": 0.3, "G": 0.3}

        total = 0.0
        for dim, score in scores.items():
            total += score * weights.get(dim, 0.33)

        return round(total, 2)

    def _get_score_rating(self, score: float) -> str:
        """获取评分等级

        Args:
            score: 得分 (0-100)

        Returns:
            str: 评级 (A+/A/B+/B/C+/C/D)
        """
        if score >= 90:
            return "A+"
        elif score >= 80:
            return "A"
        elif score >= 70:
            return "B+"
        elif score >= 60:
            return "B"
        elif score >= 50:
            return "C+"
        elif score >= 40:
            return "C"
        else:
            return "D"

    def _get_score_fill(self, score: float) -> Any:
        """根据评分获取填充颜色

        Args:
            score: 得分 (0-100)

        Returns:
            PatternFill: 填充样式
        """
        if score >= 80:
            return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif score >= 60:
            return PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def export_to_dataframe(self, metrics: Any) -> Any:
        """将 ESG 指标导出为 DataFrame

        Args:
            metrics: ESGMetrics 对象

        Returns:
            pandas.DataFrame: 包含所有指标的数据框
        """
        if not HAS_PANDAS:
            raise LibraryNotFoundError("未找到 pandas 库，请安装：pip install pandas")

        # 收集所有指标
        all_indicators = []

        for dim in ["E", "S", "G"]:
            indicators = self._get_dimension_indicators(metrics, dim)
            for name, value, unit, source, confidence in indicators:
                all_indicators.append(
                    {
                        "维度": dim,
                        "指标名称": name,
                        "指标值": value,
                        "单位": unit,
                        "数据来源": source,
                        "置信度": confidence,
                    }
                )

        return pd.DataFrame(all_indicators)

    def export_batch(
        self,
        metrics_list: List[Any],
        output_dir: Union[str, Path],
        filename_template: str = "{company}_{year}_ESG.xlsx",
    ) -> List[Path]:
        """批量导出多个 ESG 报告

        Args:
            metrics_list: ESGMetrics 对象列表
            output_dir: 输出目录
            filename_template: 文件名模板

        Returns:
            List[Path]: 导出的文件路径列表
        """
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        exported_files = []

        for metrics in metrics_list:
            filename = filename_template.format(
                company=metrics.company_name, year=metrics.year
            )
            file_path = output_path / filename

            try:
                self.export(metrics, file_path)
                exported_files.append(file_path)
            except ExcelExportError as e:
                logger.error(f"导出失败 [{metrics.company_name}]: {e}")

        return exported_files


def create_excel_exporter(config: Optional[ExportConfig] = None) -> ExcelExporter:
    """创建 Excel 导出器实例

    Args:
        config: 导出配置

    Returns:
        ExcelExporter: 导出器实例
    """
    return ExcelExporter(config)


def export_esg_report(
    metrics: Any,
    output_path: Union[str, Path],
    analysis_result: Optional[Any] = None,
    benchmark_data: Optional[Any] = None,
) -> Path:
    """便捷函数：导出 ESG 报告

    Args:
        metrics: ESGMetrics 对象
        output_path: 输出文件路径
        analysis_result: 可选的 AnalysisResult 对象
        benchmark_data: 可选的 BenchmarkData 对象

    Returns:
        Path: 导出的文件路径
    """
    exporter = ExcelExporter()
    return exporter.export(metrics, output_path, analysis_result, benchmark_data)